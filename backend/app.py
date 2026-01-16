from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "Register.xlsx"
SHEET_NAME = "Register"

# ===============================
# INIT EXCEL
# ===============================
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        ws.append([
            "ID",
            "Nama",
            "NPM",
            "Kelas",
            "Email",
            "No Telp",
            "Jurusan",
            "Divisi",
            "Pengalaman",
            "Motivasi",
            "Tanggal Daftar"
        ])

        wb.save(EXCEL_FILE)

init_excel()

# ===============================
# ROUTE: POST PENDAFTARAN
# ===============================
@app.route("/api/pendaftaran", methods=["POST"])
def pendaftaran():
    data = request.get_json()

    if not data or "pendaftaran" not in data:
        return jsonify({
            "status": "error",
            "message": "Format JSON salah"
        }), 400

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    for anggota in data["pendaftaran"]:
        required_fields = [
            "id", "nama", "npm", "kelas", "email",
            "no_telp", "jurusan", "divisi",
            "pengalaman", "motivasi", "tanggal_daftar"
        ]

        # validasi field
        for field in required_fields:
            if field not in anggota:
                return jsonify({
                    "status": "error",
                    "message": f"Field '{field}' tidak ada"
                }), 400

        ws.append([
            anggota["id"],
            anggota["nama"],
            anggota["npm"],
            anggota["kelas"],
            anggota["email"],
            anggota["no_telp"],
            anggota["jurusan"],
            anggota["divisi"],
            anggota["pengalaman"],
            anggota["motivasi"],
            anggota["tanggal_daftar"]
        ])

    wb.save(EXCEL_FILE)

    return jsonify({
        "status": "success",
        "message": "Data berhasil disimpan ke Excel"
    }), 200


# ===============================
# RUN
# ===============================
if __name__ == "__main__":
    app.run(debug=True, port=5000)
